import zipfile
import os
import re
import xml.etree.ElementTree as ET

import openpyxl
from openpyxl import load_workbook


def extract_dispimg_optimized(zip_file = "工作簿1.xlsx", output_dir = "static/temp"):
    """优化版：直接解析WPS的DISPIMG公式对应的图片，避免0KB文件"""

    os.makedirs(output_dir, exist_ok=True)
    image_cell_dict = simple_find_dispimg(zip_file)
    keys = []
    values = []
    for dictionary in image_cell_dict:
        for key, value in dictionary.items():
            keys.append(key)
            values.append(value)
    key_flag = 0

    # 从公式中提取ID
    formula = str(values)
    id_match = re.search(r'(ID_[A-F0-9]+)', formula)

    if id_match:
        image_id = id_match.group(1)  # 完整的ID_3993389D88A24F1786DB36DB73B6FED0
        print(f"提取到的图片ID: {image_id}")
    else:
        print("未能从公式中提取图片ID")
        return

    extracted_files = []

    with zipfile.ZipFile(zip_file, 'r') as z:
        file_list = z.namelist()

        # 查找media文件夹中的图片（只提取有效的图片文件）
        media_files = [f for f in file_list if f.startswith('xl/media/')]
        print(f"找到 {len(media_files)} 个媒体文件")

        valid_image_extensions = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff', '.wmf', '.emf'}

        for media_file in media_files:
            # 检查文件扩展名
            file_ext = os.path.splitext(media_file)[1].lower()
            if file_ext not in valid_image_extensions:
                print(f"跳过非图片文件: {media_file}")
                continue

            try:
                # 先检查文件大小
                file_info = z.getinfo(media_file)
                if file_info.file_size == 0:
                    print(f"跳过0字节文件: {media_file}")
                    continue

                print(f"提取图片文件: {media_file} ({file_info.file_size} bytes)")

                # 提取文件
                output_filename = f"dispimg_{os.path.basename(media_file)}"
                output_path = os.path.join(output_dir, output_filename)

                with z.open(media_file) as source, open(output_path, 'wb') as target:
                    target.write(source.read())

                # 验证提取的文件
                if os.path.getsize(output_path) > 0:
                    print(f"✓ 成功提取: {output_path} ({os.path.getsize(output_path)} bytes)")
                    extracted_files.append((keys[key_flag], output_path))
                    key_flag += 1
                else:
                    print(f"✗ 删除空文件: {output_path}")
                    os.remove(output_path)

            except Exception as e:
                print(f"✗ 提取失败 {media_file}: {e}")

        # # 查找drawing文件
        # drawing_files = [f for f in file_list if 'drawing' in f and f.endswith('.xml')]
        # print(f"\n找到 {len(drawing_files)} 个drawing文件:")
        #
        # target_media_file = None
        #
        # for draw_file in drawing_files:
        #     try:
        #         content = z.read(draw_file).decode('utf-8')
        #
        #         # 检查是否包含我们的图片ID
        #         if image_id in content:
        #             print(f"*** 在 {draw_file} 中找到图片ID! ***")
        #
        #             # 查找对应的关系文件
        #             rels_file = draw_file.replace('.xml', '.xml.rels')
        #             if rels_file in file_list:
        #                 print(f"找到关系文件: {rels_file}")
        #                 rels_content = z.read(rels_file).decode('utf-8')
        #
        #                 # 解析XML查找图片引用
        #                 ns = {'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'}
        #                 root = ET.fromstring(rels_content)
        #
        #                 for rel in root.findall('.//r:Relationship', ns):
        #                     target = rel.get('Target')
        #                     rel_type = rel.get('Type', '')
        #
        #                     # 只关注图片类型的引用
        #                     if ('image' in rel_type.lower() or
        #                             '../media/' in target or
        #                             'media/' in target):
        #
        #                         print(f"目标媒体文件: {target}")
        #                         print(f"关系类型: {rel_type}")
        #
        #                         # 构建完整路径
        #                         if target.startswith('../'):
        #                             media_path = target.replace('../', '')
        #                         elif target.startswith('..\\'):
        #                             media_path = target.replace('..\\', '')
        #                         else:
        #                             base_dir = os.path.dirname(draw_file)
        #                             media_path = os.path.normpath(os.path.join(base_dir, target))
        #
        #                         print(f"完整媒体路径: {media_path}")
        #
        #                         if media_path in file_list:
        #                             # 提取这个特定的媒体文件
        #                             file_info = z.getinfo(media_path)
        #                             if file_info.file_size > 0:
        #                                 output_filename = f"A2_{os.path.basename(media_path)}"
        #                                 output_path = os.path.join(output_dir, output_filename)
        #
        #                                 with z.open(media_path) as source, open(output_path, 'wb') as target_file:
        #                                     target_file.write(source.read())
        #
        #                                 if os.path.getsize(output_path) > 0:
        #                                     print(
        #                                         f"✓ A2图片已提取: {output_path} ({os.path.getsize(output_path)} bytes)")
        #                                     extracted_files.append(output_path)
        #                                     target_media_file = media_path
        #                                 else:
        #                                     os.remove(output_path)
        #                             else:
        #                                 print(f"跳过0字节文件: {media_path}")
        #                         else:
        #                             print(f"媒体文件不存在: {media_path}")
        #
        #     except Exception as e:
        #         print(f"解析drawing文件失败 {draw_file}: {e}")
        #         continue

        # 总结提取结果
        print(f"\n=== 提取总结 ===")
        print(f"总共提取了 {len(extracted_files)} 个有效图片文件:")
        for cell, file in extracted_files:
            size = os.path.getsize(file)
            print(f"  - {file} ({size} bytes)")

        if not extracted_files:
            print("⚠️ 没有提取到任何有效图片文件")

            # 最后手段：提取所有非零字节的媒体文件
            print("尝试提取所有有效的媒体文件...")
            for media_file in media_files:
                try:
                    file_info = z.getinfo(media_file)
                    if file_info.file_size > 0:
                        output_path = os.path.join(output_dir, f"backup_{os.path.basename(media_file)}")
                        with z.open(media_file) as source, open(output_path, 'wb') as target:
                            target.write(source.read())
                        print(f"备份提取: {output_path}")
                except:
                    pass

    return extracted_files


def simple_find_dispimg(file_path):
    """
    简化版本：只返回单元格地址作为键，内容作为值的字典列表
    """
    wb = openpyxl.load_workbook(file_path)
    results = []

    id_pattern = r'ID_[A-Za-z0-9]+'

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    cell_value_str = str(cell.value)
                    id_match = re.search(id_pattern, cell_value_str)
                    # print(f"检查单元格: {cell_value_str}")  # 调试信息

                    # 多种匹配模式，提高容错性
                    patterns = [
                        '=DISPIMG(',  # 标准格式
                        '=_xlfn.DISPIMG(',  # Excel自动添加的前缀
                        '_xlfn.DISPIMG(',  # 可能没有等号开头
                        'DISPIMG('  # 没有等号的情况
                    ]

                    # 检查是否匹配任一模式
                    if any(pattern in cell_value_str for pattern in patterns):
                        if id_match:
                            extracted_id = id_match.group(0)  # 匹配的完整字符串

                            cell_address = f"{cell.column_letter}{cell.row}"
                            # 返回键值对格式：{单元格地址: 提取的ID}
                            results.append({cell_address: extracted_id})

    wb.close()
    return results

# # 运行优化版
# file_path = '工作簿1.xlsx'
# # extract_dispimg_optimized()
# res = extract_dispimg_optimized()
# print(res)
# temp = simple_find_dispimg(file_path)
# print(temp)